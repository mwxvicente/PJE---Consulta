from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from time import sleep
import openpyxl

def iniciar_driver():
    driver = webdriver.Chrome()
    driver.get("https://pje-consulta-publica.tjmg.jus.br")
    sleep(5)
    return driver

def buscar_processo(driver, numero_oab, estado):
    # Digitar número OAB
    campo_oab = driver.find_element(By.XPATH, "//*[@id='fPP:Decoration:numeroOAB']")
    campo_oab.send_keys(numero_oab)

    # Selecionar estado
    estado_select = driver.find_element(By.XPATH, '//*[@id="fPP:Decoration:estadoComboOAB"]')
    Select(estado_select).select_by_visible_text(estado)

    # Clicar em pesquisar
    pesquisar = driver.find_element(By.XPATH, '//*[@id="fPP:searchProcessos"]')
    pesquisar.click()
    sleep(10)

def extrair_dados_processo(driver):
    processos = driver.find_elements(By.XPATH, "//b[@class='btn-block']")
    dados_processos = []

    for processo in processos:
        processo.click()
        sleep(10)
        
        janelas = driver.window_handles
        driver.switch_to.window(janelas[-1])
        driver.set_window_size(1280, 720)

        numero_processo = driver.find_element(By.XPATH, "//div[@class='col-sm-12 ']").text
        data_distribuicao = driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")[1].text
        sleep(5)

        movimentacoes = driver.find_elements(By.XPATH, '//div[@id="j_id132:processoEventoPanel_body"]//tr[contains(@class,"rich-table-row")]//td//div//div//span')
        lista_movimentacoes = [movimentacao.text for movimentacao in movimentacoes]

        dados_processos.append((numero_processo, data_distribuicao, lista_movimentacoes))
        
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

    return dados_processos

def salvar_dados_excel(dados_processos):
    workbook = openpyxl.Workbook()
    
    for numero_processo, data_distribuicao, movimentacoes in dados_processos:
        if numero_processo in workbook.sheetnames:
            pagina_processo = workbook[numero_processo]
        else:
            pagina_processo = workbook.create_sheet(numero_processo)
            pagina_processo['A1'] = "Número Processo"
            pagina_processo['B1'] = "Data Distribuição"
            pagina_processo['C1'] = "Movimentações"

        pagina_processo['A2'] = numero_processo
        pagina_processo['B2'] = data_distribuicao

        for index, movimentacao in enumerate(movimentacoes):
            pagina_processo[f'C{index + 3}'] = movimentacao

    workbook.save("dados.xlsx")

def main():
    numero_oab = input("Digite o número OAB: ")  # Solicitar número OAB ao usuário
    estado = input("Digite o estado (ex: SP): ")  # Solicitar estado ao usuário

    driver = iniciar_driver()
    buscar_processo(driver, numero_oab, estado)
    dados_processos = extrair_dados_processo(driver)
    salvar_dados_excel(dados_processos)

    driver.quit()

if __name__ == "__main__":
    main()
